from pathlib import Path

from openpyxl import load_workbook


def extract_xlsx(path: str | Path) -> dict:
    path = Path(path)

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    sheets = []
    text_parts = []

    for worksheet in workbook.worksheets:
        rows = []

        for row in worksheet.iter_rows(values_only=True):
            values = [
                str(value).strip()
                for value in row
                if value is not None
            ]

            if values:
                rows.append(values)

        if not rows:
            continue

        sheets.append(
            {
                "sheet_name": worksheet.title,
                "rows": rows,
            }
        )

        text_parts.append(
            f"Sheet: {worksheet.title}\n"
            + "\n".join(
                " | ".join(row)
                for row in rows
            )
        )

    return {
        "text": "\n\n".join(text_parts),
        "title": path.stem,
        "metadata": {
            "sheet_count": len(workbook.worksheets),
            "sheets": sheets,
        },
    }
